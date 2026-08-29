import argparse
import time
from collections import Counter
from io import BytesIO
from pathlib import Path

import openpyxl

from export_all_patient_query_tables import (
    FILTER_FIELD_COMBO_ID,
    FILTER_OPERATOR_COMBO_ID,
    FILTER_TABLE_COMBO_ID,
    QUERY_COMBO_ID,
    accept_selected_fields,
    add_field,
    cancel_fields_window,
    clear_selected_fields,
    get_all_table_names,
    get_main_window_uia,
    get_main_window_win32,
    get_value_edits,
    open_select_fields,
    save_query_changes,
    select_combo_value,
    select_fields_table,
)


SOURCE_QUERY_NAME = "OBSERVATORIO-GX"
TARGET_QUERY_NAME = "OBSERVATORIO-GX2"

BASE_FIELD_COUNT = 210
EXPECTED_FIELD_COUNT = 215
EXPECTED_FIRST_HEADER = "Visit Date"
EXPECTED_DEMOGRAPHIC_FIELD_COUNT = 24

NEW_QUERY_BUTTON_ID = 4
QUERY_NAME_EDIT_ID = 1001

FILTER_DATE = "01/01/2023"
FILTER_OPERATOR = ">"

CONTROL_DELAY = 0.5
TABLE_CHANGE_DELAY = 0.8

EXTRA_FIELDS = (
    ("GX Rest", "Time (sec)"),
    ("GX AT", "Time (sec)"),
    ("GX AT", "Ex Time (sec)"),
    ("GX VO2 Max", "Time (sec)"),
    ("GX VO2 Max", "Ex Time (sec)"),
)


def load_reference_headers(reference_path: Path) -> list[str]:
    """Lee la cabecera de una exportación válida de OBSERVATORIO-GX."""

    if not reference_path.is_file():
        raise FileNotFoundError(
            f"No existe el archivo de referencia: {reference_path}"
        )

    workbook = openpyxl.load_workbook(
        BytesIO(reference_path.read_bytes()),
        read_only=True,
        data_only=True,
    )

    if len(workbook.sheetnames) != 1:
        raise RuntimeError(
            "El archivo de referencia debe contener exactamente una hoja."
        )

    worksheet = workbook[workbook.sheetnames[0]]
    row_iterator = worksheet.iter_rows(values_only=True)

    try:
        raw_headers = list(next(row_iterator))
    except StopIteration as error:
        raise RuntimeError(
            "El archivo de referencia no contiene una cabecera."
        ) from error

    headers = [
        value.strip() if isinstance(value, str) else value
        for value in raw_headers
    ]

    if len(headers) != BASE_FIELD_COUNT:
        raise RuntimeError(
            "La referencia GX no tiene 210 columnas: "
            f"se encontraron {len(headers)}."
        )

    if any(
        not isinstance(value, str) or not value
        for value in headers
    ):
        raise RuntimeError(
            "La referencia GX contiene encabezados vacíos o no textuales."
        )

    if len(set(headers)) != BASE_FIELD_COUNT:
        raise RuntimeError(
            "La referencia GX contiene encabezados duplicados."
        )

    if headers[0] != EXPECTED_FIRST_HEADER:
        raise RuntimeError(
            f"La primera columna debe ser {EXPECTED_FIRST_HEADER!r}."
        )

    if "Pre Test Comments" not in headers:
        raise RuntimeError(
            "La referencia no contiene el encabezado "
            "'Pre Test Comments'."
        )

    return headers


def get_query_combo():
    window = get_main_window_win32()

    return window.child_window(
        control_id=QUERY_COMBO_ID,
        class_name="ThunderRT6ComboBox",
    )


def get_saved_query_names() -> list[str]:
    query_combo = get_query_combo()

    return [
        value.strip()
        for value in query_combo.item_texts()
        if value.strip()
    ]


def select_source_query() -> None:
    """Selecciona la consulta productiva únicamente para prevalidar."""

    query_names = get_saved_query_names()

    if SOURCE_QUERY_NAME not in query_names:
        raise RuntimeError(
            f"No existe la consulta fuente {SOURCE_QUERY_NAME!r}."
        )

    if TARGET_QUERY_NAME in query_names:
        raise RuntimeError(
            f"La consulta destino {TARGET_QUERY_NAME!r} ya existe. "
            "El script no la sobrescribirá."
        )

    query_combo = get_query_combo()

    select_combo_value(
        combo=query_combo,
        value=SOURCE_QUERY_NAME,
        label="el desplegable de consultas",
    )

    if query_combo.window_text().strip() != SOURCE_QUERY_NAME:
        raise RuntimeError(
            f"No quedó seleccionada la consulta {SOURCE_QUERY_NAME!r}."
        )


def resolve_field_name(
    table_name: str,
    exported_name: str,
    available_fields: list[str],
) -> str:
    """Resuelve el nombre visible del campo dentro de Patient Query."""

    exact_matches = [
        field_name
        for field_name in available_fields
        if field_name == exported_name
    ]

    if len(exact_matches) == 1:
        return exact_matches[0]

    qualified_matches = [
        field_name
        for field_name in available_fields
        if field_name.startswith(f"{exported_name} [")
    ]

    if len(qualified_matches) == 1:
        return qualified_matches[0]

    candidates = exact_matches or qualified_matches

    raise RuntimeError(
        f"No se pudo resolver de forma única "
        f"{table_name} > {exported_name!r}; "
        f"coincidencias={candidates!r}."
    )


def build_field_plan(
    headers: list[str],
) -> list[tuple[str, str]]:
    """
    Construye el plan con los 210 campos actuales y los cinco nuevos.
    Valida todo antes de crear la consulta.
    """

    table_names = get_all_table_names()

    if "Demographic" not in table_names:
        raise RuntimeError(
            "No se encontró la tabla 'Demographic'."
        )

    demographic_fields = select_fields_table("Demographic")

    gx_pf_tables = sorted(
        [
            table_name
            for table_name in table_names
            if (
                table_name.startswith("GX ")
                or table_name.startswith("PF ")
            )
        ],
        key=len,
        reverse=True,
    )

    requested: list[tuple[str, str]] = []

    for header in headers:
        if header in demographic_fields:
            requested.append(("Demographic", header))
            continue

        matching_tables = [
            table_name
            for table_name in gx_pf_tables
            if header.startswith(f"{table_name} ")
        ]

        if not matching_tables:
            raise RuntimeError(
                "No se pudo identificar la tabla del encabezado "
                f"{header!r}."
            )

        table_name = matching_tables[0]
        exported_name = header[len(table_name) + 1 :]

        requested.append(
            (table_name, exported_name)
        )

    requested.extend(EXTRA_FIELDS)

    requested_counts = Counter(
        table_name
        for table_name, _ in requested
    )

    if (
        requested_counts["Demographic"]
        != EXPECTED_DEMOGRAPHIC_FIELD_COUNT
    ):
        raise RuntimeError(
            "La referencia no produjo los 24 campos "
            "Demographic esperados: "
            f"se resolvieron "
            f"{requested_counts['Demographic']}."
        )

    available_by_table: dict[str, list[str]] = {
        "Demographic": demographic_fields,
    }

    for table_name in dict.fromkeys(
        table_name
        for table_name, _ in requested
    ):
        if table_name == "Demographic":
            continue

        available_by_table[table_name] = (
            select_fields_table(table_name)
        )

    resolved = [
        (
            table_name,
            resolve_field_name(
                table_name,
                exported_name,
                available_by_table[table_name],
            ),
        )
        for table_name, exported_name in requested
    ]

    if len(resolved) != EXPECTED_FIELD_COUNT:
        raise RuntimeError(
            f"El plan produjo {len(resolved)} campos, "
            f"no {EXPECTED_FIELD_COUNT}."
        )

    if len(set(resolved)) != EXPECTED_FIELD_COUNT:
        duplicates = [
            field
            for field, count in Counter(resolved).items()
            if count > 1
        ]

        raise RuntimeError(
            "El plan contiene campos duplicados: "
            f"{duplicates!r}."
        )

    return resolved


def print_plan_summary(
    field_plan: list[tuple[str, str]],
) -> None:
    counts = Counter(
        table_name
        for table_name, _ in field_plan
    )

    print(
        f"Prevalidación correcta: "
        f"{len(field_plan)} campos resueltos."
    )

    for table_name, count in counts.items():
        print(f"  {table_name}: {count}")

    print("Campos nuevos:")

    for table_name, field_name in EXTRA_FIELDS:
        print(f"  {table_name} > {field_name}")


def create_new_query_shell() -> None:
    """Pulsa Nuevo y asigna el nombre OBSERVATORIO-GX2."""

    query_names = get_saved_query_names()

    if TARGET_QUERY_NAME in query_names:
        raise RuntimeError(
            f"La consulta {TARGET_QUERY_NAME!r} ya existe. "
            "No se sobrescribirá."
        )

    window = get_main_window_win32()

    new_button = window.child_window(
        control_id=NEW_QUERY_BUTTON_ID,
        class_name="ThunderRT6CommandButton",
    )

    new_button.wait(
        "visible enabled",
        timeout=30,
    )

    new_button.click_input()
    time.sleep(CONTROL_DELAY)

    query_combo = get_query_combo()

    query_name_edit = query_combo.child_window(
        control_id=QUERY_NAME_EDIT_ID,
    )

    query_name_edit.wait(
        "visible enabled",
        timeout=10,
    )

    query_name_edit.set_edit_text(
        TARGET_QUERY_NAME
    )

    time.sleep(CONTROL_DELAY)

    current_name = query_name_edit.window_text().strip()

    if current_name != TARGET_QUERY_NAME:
        raise RuntimeError(
            "El nombre de la consulta nueva no quedó configurado. "
            f"Esperado={TARGET_QUERY_NAME!r}; "
            f"actual={current_name!r}."
        )

    print(
        f"Consulta nueva preparada: {TARGET_QUERY_NAME}"
    )


def configure_first_filter() -> None:
    """
    Configura la primera fila:

    Demographic > Visit Date > 01/01/2023
    """

    window_win32 = get_main_window_win32()

    table_combo = window_win32.child_window(
        control_id=FILTER_TABLE_COMBO_ID,
        class_name="ThunderRT6ComboBox",
    )

    field_combo = window_win32.child_window(
        control_id=FILTER_FIELD_COMBO_ID,
        class_name="ThunderRT6ComboBox",
    )

    operator_combo = window_win32.child_window(
        control_id=FILTER_OPERATOR_COMBO_ID,
        class_name="ThunderRT6ComboBox",
    )

    select_combo_value(
        combo=table_combo,
        value="Demographic",
        label="la tabla de la primera fila del filtro",
    )

    time.sleep(TABLE_CHANGE_DELAY)

    select_combo_value(
        combo=field_combo,
        value="Visit Date",
        label="el campo de la primera fila del filtro",
    )

    time.sleep(CONTROL_DELAY)

    select_combo_value(
        combo=operator_combo,
        value=FILTER_OPERATOR,
        label="el operador de la primera fila del filtro",
    )

    time.sleep(CONTROL_DELAY)

    window_uia = get_main_window_uia()
    value_edits = get_value_edits(window_uia)

    if not value_edits:
        raise RuntimeError(
            "No se encontraron campos de valor visibles."
        )

    value_edit = value_edits[0]

    value_edit.set_text(FILTER_DATE)
    value_edit.click_input()
    value_edit.type_keys("{TAB}")

    time.sleep(CONTROL_DELAY)

    try:
        current_value = value_edit.get_value()
    except Exception:
        current_value = value_edit.window_text()

    if current_value != FILTER_DATE:
        raise RuntimeError(
            "La fecha del filtro no quedó configurada. "
            f"Esperada={FILTER_DATE!r}; "
            f"actual={current_value!r}."
        )

    print(
        "Filtro configurado: "
        f"Demographic > Visit Date "
        f"{FILTER_OPERATOR} {FILTER_DATE}"
    )


def populate_new_query(
    field_plan: list[tuple[str, str]],
) -> None:
    """Añade los 215 campos a la nueva consulta."""

    open_select_fields()
    fields_window_open = True

    try:
        clear_selected_fields()

        total = len(field_plan)

        for position, (
            table_name,
            field_name,
        ) in enumerate(
            field_plan,
            start=1,
        ):
            print(
                f"[{position:03d}/{total}] "
                f"{table_name} > {field_name}"
            )

            add_field(
                table_name=table_name,
                field_name=field_name,
            )

        accept_selected_fields()
        fields_window_open = False

    finally:
        if fields_window_open:
            try:
                cancel_fields_window()
            except Exception:
                pass


def verify_saved_query() -> None:
    """Confirma que la consulta nueva aparece en el desplegable."""

    query_names = get_saved_query_names()

    if TARGET_QUERY_NAME not in query_names:
        raise RuntimeError(
            f"La consulta {TARGET_QUERY_NAME!r} "
            "no aparece entre las consultas guardadas."
        )

    query_combo = get_query_combo()

    if query_combo.window_text().strip() != TARGET_QUERY_NAME:
        select_combo_value(
            combo=query_combo,
            value=TARGET_QUERY_NAME,
            label="el desplegable de consultas",
        )

    if query_combo.window_text().strip() != TARGET_QUERY_NAME:
        raise RuntimeError(
            "No se pudo verificar la consulta guardada."
        )

    print(
        f"Consulta guardada y verificada: "
        f"{TARGET_QUERY_NAME}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Crea OBSERVATORIO-GX2 con los 210 campos "
            "de una exportación GX válida y cinco campos "
            "temporales adicionales en segundos."
        )
    )

    parser.add_argument(
        "--reference",
        required=True,
        type=Path,
        help=(
            "Archivo XLS válido exportado por "
            "OBSERVATORIO-GX con 210 columnas."
        ),
    )

    parser.add_argument(
        "--apply",
        action="store_true",
        help=(
            "Crea y guarda OBSERVATORIO-GX2. "
            "Sin esta opción solo realiza la prevalidación."
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    reference_path = args.reference.resolve()
    headers = load_reference_headers(reference_path)

    print(f"Referencia: {reference_path}")
    print(f"Campos base: {len(headers)}")

    select_source_query()

    fields_window_open = False
    target_started = False

    try:
        open_select_fields()
        fields_window_open = True

        field_plan = build_field_plan(headers)
        print_plan_summary(field_plan)

        cancel_fields_window()
        fields_window_open = False

        if not args.apply:
            print(
                "Modo validación: no se creó ni modificó "
                "ninguna consulta."
            )
            return

        create_new_query_shell()
        target_started = True

        populate_new_query(field_plan)
        configure_first_filter()
        save_query_changes()
        verify_saved_query()

        print(
            f"Proceso completado: "
            f"{TARGET_QUERY_NAME} tiene "
            f"{EXPECTED_FIELD_COUNT} campos."
        )

    except Exception:
        if fields_window_open:
            try:
                cancel_fields_window()
            except Exception:
                pass

        if target_started:
            print(
                "La creación de la consulta se interrumpió. "
                "No pulses Guardar manualmente; revisa el error."
            )

        raise


if __name__ == "__main__":
    main()