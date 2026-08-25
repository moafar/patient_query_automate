import argparse
from collections import Counter
from io import BytesIO
from pathlib import Path

import openpyxl

from export_all_patient_query_tables import (
    QUERY_COMBO_ID,
    add_field,
    accept_selected_fields,
    cancel_fields_window,
    clear_selected_fields,
    get_all_table_names,
    get_main_window_win32,
    open_select_fields,
    save_query_changes,
    select_combo_value,
    select_fields_table,
)


QUERY_NAME = "OBSERVATORIO-GX"
EXPECTED_FIELD_COUNT = 210
EXPECTED_FIRST_HEADER = "Visit Date"
EXPECTED_DEMOGRAPHIC_FIELD_COUNT = 24


def load_reference_headers(reference_path: Path) -> list[str]:
    """Lee únicamente la cabecera de una exportación GX INO válida."""
    if not reference_path.is_file():
        raise FileNotFoundError(
            f"No existe el archivo de referencia: {reference_path}"
        )

    workbook = openpyxl.load_workbook(
        BytesIO(reference_path.read_bytes()),
        read_only=True,
        data_only=True,
    )

    if len(workbook.sheetnames) != 1:
        raise RuntimeError(
            "El archivo de referencia debe contener exactamente una hoja."
        )

    worksheet = workbook[workbook.sheetnames[0]]
    row_iterator = worksheet.iter_rows(values_only=True)

    try:
        raw_headers = list(next(row_iterator))
    except StopIteration as error:
        raise RuntimeError(
            "El archivo de referencia no contiene una cabecera."
        ) from error

    headers = [
        value.strip() if isinstance(value, str) else value
        for value in raw_headers
    ]

    if len(headers) != EXPECTED_FIELD_COUNT:
        raise RuntimeError(
            "La referencia GX INO no tiene 210 columnas: "
            f"se encontraron {len(headers)}."
        )

    if any(not isinstance(value, str) or not value for value in headers):
        raise RuntimeError(
            "La referencia GX INO contiene encabezados vacíos o no textuales."
        )

    if len(set(headers)) != EXPECTED_FIELD_COUNT:
        raise RuntimeError(
            "La referencia GX INO contiene encabezados duplicados."
        )

    if headers[0] != EXPECTED_FIRST_HEADER:
        raise RuntimeError(
            f"La primera columna debe ser {EXPECTED_FIRST_HEADER!r}."
        )

    if "Pre Test Comments" not in headers:
        raise RuntimeError(
            "La referencia no contiene el encabezado 'Pre Test Comments'."
        )

    return headers


def select_existing_query() -> None:
    """Selecciona GX INO de forma estricta; nunca escribe un nombre alterno."""
    window = get_main_window_win32()
    query_combo = window.child_window(
        control_id=QUERY_COMBO_ID,
        class_name="ThunderRT6ComboBox",
    )

    select_combo_value(
        combo=query_combo,
        value=QUERY_NAME,
        label="el desplegable de consultas",
    )

    if query_combo.window_text() != QUERY_NAME:
        raise RuntimeError(
            f"No quedó seleccionada exactamente la consulta {QUERY_NAME!r}."
        )


def resolve_field_name(
    table_name: str,
    exported_name: str,
    available_fields: list[str],
) -> str:
    """Resuelve el nombre visible, incluidos sufijos como [FVLData]."""
    exact_matches = [
        field_name
        for field_name in available_fields
        if field_name == exported_name
    ]
    if len(exact_matches) == 1:
        return exact_matches[0]

    qualified_matches = [
        field_name
        for field_name in available_fields
        if field_name.startswith(f"{exported_name} [")
    ]
    if len(qualified_matches) == 1:
        return qualified_matches[0]

    candidates = exact_matches or qualified_matches
    raise RuntimeError(
        f"No se pudo resolver de forma única {table_name} > "
        f"{exported_name!r}; coincidencias={candidates!r}."
    )


def build_field_plan(headers: list[str]) -> list[tuple[str, str]]:
    """Valida todo el catálogo vivo antes de modificar la consulta."""
    table_names = get_all_table_names()
    if "Demographic" not in table_names:
        raise RuntimeError("No se encontró la tabla 'Demographic'.")

    demographic_fields = select_fields_table("Demographic")

    gx_pf_tables = sorted(
        [
            table_name
            for table_name in table_names
            if table_name.startswith("GX ") or table_name.startswith("PF ")
        ],
        key=len,
        reverse=True,
    )

    requested: list[tuple[str, str]] = []
    for header in headers:
        if header in demographic_fields:
            requested.append(("Demographic", header))
            continue

        matching_tables = [
            table_name
            for table_name in gx_pf_tables
            if header.startswith(f"{table_name} ")
        ]
        if not matching_tables:
            raise RuntimeError(
                f"No se pudo identificar la tabla del encabezado {header!r}."
            )

        table_name = matching_tables[0]
        exported_name = header[len(table_name) + 1 :]
        requested.append((table_name, exported_name))

    requested_counts = Counter(table_name for table_name, _ in requested)
    if requested_counts["Demographic"] != EXPECTED_DEMOGRAPHIC_FIELD_COUNT:
        raise RuntimeError(
            "La referencia no produjo los 24 campos Demographic esperados: "
            f"se resolvieron {requested_counts['Demographic']}."
        )

    available_by_table: dict[str, list[str]] = {
        "Demographic": demographic_fields,
    }
    for table_name in dict.fromkeys(table for table, _ in requested):
        if table_name == "Demographic":
            continue
        available_by_table[table_name] = select_fields_table(table_name)

    resolved = [
        (
            table_name,
            resolve_field_name(
                table_name,
                exported_name,
                available_by_table[table_name],
            ),
        )
        for table_name, exported_name in requested
    ]

    if len(resolved) != EXPECTED_FIELD_COUNT:
        raise RuntimeError(
            f"El plan produjo {len(resolved)} campos, no 210."
        )

    if len(set(resolved)) != EXPECTED_FIELD_COUNT:
        raise RuntimeError("El plan contiene campos duplicados.")

    return resolved


def print_plan_summary(field_plan: list[tuple[str, str]]) -> None:
    counts = Counter(table_name for table_name, _ in field_plan)
    print("Prevalidación correcta: 210 campos resueltos.")
    for table_name, count in counts.items():
        print(f"  {table_name}: {count}")


def rebuild_query(field_plan: list[tuple[str, str]]) -> None:
    clear_selected_fields()

    total = len(field_plan)
    for position, (table_name, field_name) in enumerate(field_plan, start=1):
        print(f"[{position:03d}/{total}] {table_name} > {field_name}")
        add_field(table_name=table_name, field_name=field_name)

    accept_selected_fields()
    save_query_changes()
    print("Consulta GX INO reconstruida y guardada con 210 campos.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Prevalida o reconstruye los 210 campos de GX INO usando "
            "la cabecera de una exportación válida."
        )
    )
    parser.add_argument(
        "--reference",
        required=True,
        type=Path,
        help="Archivo GX INO válido de 210 columnas.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Aplica y guarda la reconstrucción; sin esta opción solo valida.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    headers = load_reference_headers(args.reference.resolve())
    select_existing_query()

    fields_window_open = False
    try:
        open_select_fields()
        fields_window_open = True
        field_plan = build_field_plan(headers)
        print_plan_summary(field_plan)

        if not args.apply:
            cancel_fields_window()
            fields_window_open = False
            print("Modo validación: no se modificó la consulta.")
            return

        rebuild_query(field_plan)
        fields_window_open = False
    except Exception:
        if fields_window_open:
            try:
                cancel_fields_window()
            except Exception:
                pass
        raise


if __name__ == "__main__":
    main()
